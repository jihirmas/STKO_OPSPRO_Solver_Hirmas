"""
assignment_registry_dialog.py
------------------------------
QDialog that displays the contents of an AssignmentRegistry in a
sortable / filterable QTableView.

Columns
-------
0  Component Name  str    component.name
1  Component ID    int    component.id
2  Component Group str    displayName of the component's group
3  Target Type     str    "Geometry" | "Interaction"
4  Target ID       int    geometry.id  | interaction.id
5  Subshape Type   str    "Vertex" | "Edge" | "Face" | "Solid" | "" (Interaction)
6  Subshape ID     int    subshape index  |  "" (Interaction)
"""

from PySide2 import QtCore, QtGui, QtWidgets
from PyMpc import App, MpcSubshapeType
import os
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

from opspro.utils.assignment_registry import AssignmentRegistry
from opspro.utils.file_dialog_manager import FileDialogManager


# ---------------------------------------------------------------------------
# Column indices and headers
# ---------------------------------------------------------------------------

_COL_COMP_NAME   = 0
_COL_COMP_ID     = 1
_COL_COMP_GROUP  = 2
_COL_TARGET_TYPE = 3
_COL_TARGET_ID   = 4
_COL_SUB_TYPE    = 5
_COL_SUB_ID      = 6

_HEADERS = [
    'Component Name',
    'Component ID',
    'Component Group',
    'Target Type',
    'Target ID',
    'Subshape Type',
    'Subshape ID',
]

_SUBSHAPE_TYPES = [
    (MpcSubshapeType.Vertex, 'Vertex'),
    (MpcSubshapeType.Edge,   'Edge'),
    (MpcSubshapeType.Face,   'Face'),
    (MpcSubshapeType.Solid,  'Solid'),
]


# ---------------------------------------------------------------------------
# Row builder
# ---------------------------------------------------------------------------

def _build_rows(registry: AssignmentRegistry) -> list:
    """
    Flatten the AssignmentRegistry into a list of 7-tuples::

        (comp_name, comp_id, comp_group, target_type, target_id, sub_type, sub_id)

    Subshape columns are empty strings for Interaction rows.
    """
    # Build group_id -> displayName mapping once from the active document.
    # Falls back to the raw group ID string if the document is unavailable.
    _group_display = {}
    try:
        doc = App.caeDocument()
        if doc is not None:
            groups = doc.pluginCaeComponents.groups()
            for gid in groups.keys():
                _group_display[gid] = groups[gid].displayName
    except Exception:
        pass

    def _group_name(comp) -> str:
        gid = comp.componentGroupID()
        return _group_display.get(gid, str(gid))
    rows = []

    # --- Geometry assignments -----------------------------------------
    for geometry, item in registry.geometry_assignment.items():
        geom_id = int(geometry.id)
        for subshape_type, subshape_name in _SUBSHAPE_TYPES:
            sub = item.get_by_subshape_type(subshape_type)
            for subshape_id, components in sub.items():
                for comp in components:
                    rows.append((
                        str(comp.name),
                        int(comp.id),
                        _group_name(comp),
                        'Geometry',
                        geom_id,
                        subshape_name,
                        int(subshape_id),
                    ))

    # --- Interaction assignments ---------------------------------------
    for interaction, components in registry.interaction_assignment.items():
        inter_id = int(interaction.id)
        for comp in components:
            rows.append((
                str(comp.name),
                int(comp.id),
                _group_name(comp),
                'Interaction',
                inter_id,
                '',
                '',
            ))

    return rows


# ---------------------------------------------------------------------------
# Table model
# ---------------------------------------------------------------------------

class AssignmentTableModel(QtCore.QAbstractTableModel):
    """
    Read-only table model backed by a flat list of 7-tuples produced by
    :func:`_build_rows`.

    Geometry rows have all seven columns populated; Interaction rows have
    empty strings in the Subshape Type and Subshape ID columns.
    """

    def __init__(self, rows: list, parent=None):
        super().__init__(parent)
        self._rows = rows
        # Pre-computed lowercase concatenation of every cell — used by the
        # free-text search in the proxy to avoid per-cell model.data() calls.
        self._search_strings = [
            '\t'.join(str(v).lower() for v in row)
            for row in rows
        ]

    # ------------------------------------------------------------------
    # Qt API
    # ------------------------------------------------------------------

    def rowCount(self, parent=QtCore.QModelIndex()):
        return 0 if parent.isValid() else len(self._rows)

    def columnCount(self, parent=QtCore.QModelIndex()):
        return 0 if parent.isValid() else len(_HEADERS)

    def headerData(self, section, orientation, role=QtCore.Qt.DisplayRole):
        if role == QtCore.Qt.DisplayRole and orientation == QtCore.Qt.Horizontal:
            return _HEADERS[section]
        return None

    def data(self, index, role=QtCore.Qt.DisplayRole):
        if not index.isValid():
            return None

        row = self._rows[index.row()]
        col = index.column()
        value = row[col]

        if role == QtCore.Qt.DisplayRole:
            return str(value)

        if role == QtCore.Qt.TextAlignmentRole:
            if col in (_COL_COMP_ID, _COL_TARGET_ID, _COL_SUB_ID):
                return int(QtCore.Qt.AlignCenter)
            return int(QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)

        if role == QtCore.Qt.ForegroundRole:
            # Grey out the subshape columns for Interaction rows
            if col in (_COL_SUB_TYPE, _COL_SUB_ID) and row[_COL_TARGET_TYPE] == 'Interaction':
                return QtGui.QColor(160, 160, 160)

        return None

    def flags(self, index):
        return QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable


# ---------------------------------------------------------------------------
# Custom sort/filter proxy
# ---------------------------------------------------------------------------

class _AssignmentFilterProxy(QtCore.QSortFilterProxyModel):
    """
    Proxy that supports two independent filter conditions applied together:

    * Free-text search across all columns (case-insensitive substring).
    * Exact match on Target Type column ("Geometry" | "Interaction" | "").
    * Numeric sort for the integer columns (Component ID, Target ID, Subshape ID).
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self._text_filter      = ''   # lower-cased search string
        self._target_filter    = ''   # '' means "All"
        self._comp_name_filter = ''   # '' means "All"
        self._comp_group_filter = ''  # '' means "All"
        self.setSortCaseSensitivity(QtCore.Qt.CaseInsensitive)

    def set_text(self, text: str):
        self._text_filter = text.strip().lower()
        self.invalidateFilter()

    def set_target_type(self, label: str):
        self._target_filter = label if label != 'All' else ''
        self.invalidateFilter()

    def set_comp_name(self, label: str):
        self._comp_name_filter = label if label != 'All' else ''
        self.invalidateFilter()

    def set_comp_group(self, label: str):
        self._comp_group_filter = label if label != 'All' else ''
        self.invalidateFilter()

    # ------------------------------------------------------------------
    # Overrides
    # ------------------------------------------------------------------

    def filterAcceptsRow(self, source_row: int, source_parent: QtCore.QModelIndex) -> bool:
        # Access the raw tuple directly — avoids Python→C++ model.data() dispatch
        # for every cell of every row on each filter invalidation.
        src = self.sourceModel()
        row = src._rows[source_row]

        # --- target type exact filter ---------------------------------
        if self._target_filter and row[_COL_TARGET_TYPE] != self._target_filter:
            return False

        # --- component name exact filter ------------------------------
        if self._comp_name_filter and str(row[_COL_COMP_NAME]) != self._comp_name_filter:
            return False

        # --- component group exact filter -----------------------------
        if self._comp_group_filter and str(row[_COL_COMP_GROUP]) != self._comp_group_filter:
            return False

        # --- free-text search — use pre-computed lowercase string -----
        if self._text_filter:
            return self._text_filter in src._search_strings[source_row]

        return True

    def lessThan(self, left: QtCore.QModelIndex, right: QtCore.QModelIndex) -> bool:
        col = left.column()
        if col in (_COL_COMP_ID, _COL_TARGET_ID, _COL_SUB_ID):
            # Access _rows directly to avoid model.data() dispatch overhead
            rows = self.sourceModel()._rows
            l_val = rows[left.row()][col]
            r_val = rows[right.row()][col]
            if isinstance(l_val, int) and isinstance(r_val, int):
                return l_val < r_val
            try:
                return int(l_val) < int(r_val)
            except (ValueError, TypeError):
                pass
        return super().lessThan(left, right)


# ---------------------------------------------------------------------------
# Dialog
# ---------------------------------------------------------------------------

class AssignmentRegistryDialog(QtWidgets.QDialog):
    """
    Read-only dialog that displays all entries of an
    :class:`~opspro.utils.assignment_registry.AssignmentRegistry`.

    The table is sortable by clicking column headers and filterable via
    the search box and the Target Type combo at the top.

    Parameters
    ----------
    registry : AssignmentRegistry
        Registry instance whose data should be displayed.
    parent : QWidget, optional
    """

    def __init__(self, registry: AssignmentRegistry, parent=None):
        super().__init__(parent)
        try:
            self.setWindowTitle('Assignment Registry')
            self.resize(860, 500)
            self._registry = registry
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, 'Error', f'Failed to initialize dialog:\n{e}')
            self.close()
            return
        self._setup_ui()
        self._load_data()

    # ------------------------------------------------------------------
    # UI construction
    # ------------------------------------------------------------------

    def _setup_ui(self):
        try:
            # ---- source model + proxy ------------------------------------
            self._source_model = AssignmentTableModel([], self)
            self._proxy = _AssignmentFilterProxy(self)
            self._proxy.setSourceModel(self._source_model)

            # ---- description label ---------------------------------------
            lbl_desc = QtWidgets.QLabel(
                'This table shows all active assignments between CAE components and their targets.\n'
                'Use <b>Search</b> to find entries across any column, or the <b>Filter by</b> combos '
                'to narrow results by target type, component name, or component group.\n'
                'Click a column header to sort. Use <b>Export to Excel</b> to export the current view.'
            )
            lbl_desc.setWordWrap(True)
            lbl_desc.setTextFormat(QtCore.Qt.RichText)

            # ---- search + filter grid (column 0 = labels, aligned) ------
            top_grid = QtWidgets.QGridLayout()
            top_grid.setSpacing(6)
            top_grid.setColumnStretch(1, 1)

            # row 0: Search
            top_grid.addWidget(QtWidgets.QLabel('Search:'), 0, 0)
            self._search_edit = QtWidgets.QLineEdit()
            self._search_edit.setPlaceholderText('Filter across all columns…')
            self._search_edit.setClearButtonEnabled(True)
            self._search_edit.textChanged.connect(self._on_search_text_changed)
            top_grid.addWidget(self._search_edit, 0, 1)

            # debounce timer: apply text filter 200 ms after the user stops typing
            self._search_timer = QtCore.QTimer(self)
            self._search_timer.setSingleShot(True)
            self._search_timer.setInterval(200)
            self._search_timer.timeout.connect(self._on_search_changed)

            # row 1: Filter by …
            top_grid.addWidget(QtWidgets.QLabel('Filter by:'), 1, 0)
            filter_inner = QtWidgets.QHBoxLayout()
            filter_inner.setSpacing(6)
            filter_inner.addWidget(QtWidgets.QLabel('Target:'))
            self._combo_target = QtWidgets.QComboBox()
            self._combo_target.addItems(['All', 'Geometry', 'Interaction'])
            self._combo_target.setMinimumWidth(100)
            self._combo_target.currentTextChanged.connect(self._on_target_filter_changed)
            filter_inner.addWidget(self._combo_target)
            filter_inner.addSpacing(8)
            filter_inner.addWidget(QtWidgets.QLabel('Component Name:'))
            self._combo_comp_name = QtWidgets.QComboBox()
            self._combo_comp_name.setMinimumWidth(120)
            self._combo_comp_name.currentTextChanged.connect(self._on_comp_name_filter_changed)
            filter_inner.addWidget(self._combo_comp_name)
            filter_inner.addSpacing(8)
            filter_inner.addWidget(QtWidgets.QLabel('Component Group:'))
            self._combo_comp_group = QtWidgets.QComboBox()
            self._combo_comp_group.setMinimumWidth(120)
            self._combo_comp_group.currentTextChanged.connect(self._on_comp_group_filter_changed)
            filter_inner.addWidget(self._combo_comp_group)
            filter_inner.addStretch()
            top_grid.addLayout(filter_inner, 1, 1)

            # ---- table view ----------------------------------------------
            self._table = QtWidgets.QTableView()
            self._table.setModel(self._proxy)
            self._table.setSortingEnabled(True)
            self._table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
            self._table.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
            self._table.setAlternatingRowColors(True)
            self._table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
            self._table.setVerticalScrollMode(QtWidgets.QAbstractItemView.ScrollPerPixel)
            self._table.setHorizontalScrollMode(QtWidgets.QAbstractItemView.ScrollPerPixel)
            vh = self._table.verticalHeader()
            vh.hide()
            vh.setDefaultSectionSize(22)  # fixed row height \u2014 avoids per-row measurement during paint

            hh = self._table.horizontalHeader()
            hh.setSectionsMovable(True)
            hh.setStretchLastSection(False)
            # Interactive resize: Qt never scans all rows to compute widths.
            # Initial widths are set after data is loaded in _set_initial_column_widths().
            hh.setSectionResizeMode(QtWidgets.QHeaderView.Interactive)
            hh.setSectionResizeMode(_COL_COMP_NAME, QtWidgets.QHeaderView.Stretch)

            # default sort: Component ID ascending
            self._proxy.sort(_COL_COMP_ID, QtCore.Qt.AscendingOrder)

            # ---- row count label -----------------------------------------
            self._lbl_count = QtWidgets.QLabel()

            # ---- separator -----------------------------------------------
            separator = QtWidgets.QFrame()
            separator.setFrameShape(QtWidgets.QFrame.HLine)
            separator.setFrameShadow(QtWidgets.QFrame.Sunken)

            # ---- button box ----------------------------------------------
            btn_box = QtWidgets.QDialogButtonBox(
                QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel,
                QtCore.Qt.Horizontal,
            )
            btn_box.accepted.connect(self.accept)
            btn_box.rejected.connect(self.reject)

            # Export button on the left side of the button box
            btn_export = btn_box.addButton('Export to Excel…', QtWidgets.QDialogButtonBox.ActionRole)
            btn_export.clicked.connect(self._on_export_excel)

            # ---- main layout --------------------------------------------
            vbox = QtWidgets.QVBoxLayout(self)
            vbox.setContentsMargins(8, 8, 8, 8)
            vbox.setSpacing(6)
            vbox.addWidget(lbl_desc)
            vbox.addLayout(top_grid)
            vbox.addWidget(self._table)
            vbox.addWidget(self._lbl_count)
            vbox.addWidget(separator)
            vbox.addWidget(btn_box)

        except Exception as e:
            QtWidgets.QMessageBox.critical(self, 'Error', f'Failed to set up UI:\n{e}')
            self.close()

    # ------------------------------------------------------------------
    # Data loading
    # ------------------------------------------------------------------

    def _load_data(self):
        try:
            rows = _build_rows(self._registry)
            self._source_model = AssignmentTableModel(rows, self)
            self._proxy.setSourceModel(self._source_model)
            self._proxy.sort(_COL_COMP_ID, QtCore.Qt.AscendingOrder)
            self._populate_filter_combos(rows)
            self._set_initial_column_widths()
            self._update_count()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, 'Error', f'Failed to load data:\n{e}')
            self.close()

    # Fixed pixel widths for non-stretch columns.
    # Wide enough for typical content; user can drag to resize.
    _COLUMN_WIDTHS = {
        _COL_COMP_ID:     90,
        _COL_COMP_GROUP: 130,
        _COL_TARGET_TYPE: 90,
        _COL_TARGET_ID:   80,
        _COL_SUB_TYPE:    90,
        _COL_SUB_ID:      80,
    }

    def _set_initial_column_widths(self):
        hh = self._table.horizontalHeader()
        for col, width in self._COLUMN_WIDTHS.items():
            hh.resizeSection(col, width)

    def _populate_filter_combos(self, rows: list):
        """Fill component-name and component-group combos with sorted unique values."""
        names  = sorted({r[_COL_COMP_NAME]  for r in rows})
        groups = sorted({r[_COL_COMP_GROUP] for r in rows})
        for combo, values in (
            (self._combo_comp_name,  names),
            (self._combo_comp_group, groups),
        ):
            combo.blockSignals(True)
            combo.clear()
            combo.addItem('All')
            combo.addItems(values)
            combo.blockSignals(False)

    # ------------------------------------------------------------------
    # Slots
    # ------------------------------------------------------------------

    def _on_search_text_changed(self, text: str):
        """Restart the debounce timer on every keystroke."""
        self._search_timer.start()

    def _on_search_changed(self):
        try:
            self._proxy.set_text(self._search_edit.text())
            self._update_count()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, 'Error', f'Failed to apply search filter:\n{e}')

    def _on_target_filter_changed(self, text: str):
        try:
            self._proxy.set_target_type(text)
            self._update_count()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, 'Error', f'Failed to apply target type filter:\n{e}')

    def _on_comp_name_filter_changed(self, text: str):
        try:
            self._proxy.set_comp_name(text)
            self._update_count()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, 'Error', f'Failed to apply component name filter:\n{e}')

    def _on_comp_group_filter_changed(self, text: str):
        try:
            self._proxy.set_comp_group(text)
            self._update_count()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, 'Error', f'Failed to apply component group filter:\n{e}')

    def _update_count(self):
        try:
            total   = self._source_model.rowCount()
            visible = self._proxy.rowCount()
            self._lbl_count.setText(f'{visible} / {total} rows')
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, 'Error', f'Failed to update row count:\n{e}')

    def _on_export_excel(self):
        if not _OPENPYXL_AVAILABLE:
            QtWidgets.QMessageBox.warning(
                self, 'Missing dependency',
                'openpyxl is not installed.\n\nInstall it with:\n    pip install openpyxl'
            )
            return

        FileDialogManager.set_context('LD_ExportAssignmentsExcel')
        path = FileDialogManager.get_save_file_name(
            self, 'Export to Excel',
            'Excel workbook (*.xlsx)'
        )
        if not path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Assignments'

            # ---- header row ------------------------------------------
            header_font  = Font(bold=True, color='FFFFFF')
            header_fill  = PatternFill('solid', fgColor='366092')
            header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

            for col_idx, label in enumerate(_HEADERS, start=1):
                cell = ws.cell(row=1, column=col_idx, value=label)
                cell.font  = header_font
                cell.fill  = header_fill
                cell.alignment = header_align

            # ---- data rows (respects current filter + sort order) ----
            alt_fill = PatternFill('solid', fgColor='DCE6F1')
            for row_idx in range(self._proxy.rowCount()):
                excel_row = row_idx + 2
                for col_idx in range(self._proxy.columnCount()):
                    proxy_idx = self._proxy.index(row_idx, col_idx)
                    value = self._proxy.data(proxy_idx, QtCore.Qt.DisplayRole)
                    cell = ws.cell(row=excel_row, column=col_idx + 1, value=value)
                    if row_idx % 2 == 1:
                        cell.fill = alt_fill

            # ---- column widths (approximate) -------------------------
            for col_idx, label in enumerate(_HEADERS, start=1):
                ws.column_dimensions[
                    openpyxl.utils.get_column_letter(col_idx)
                ].width = max(12, len(label) + 4)

            ws.freeze_panes = 'A2'
            wb.save(path)
            os.startfile(path)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, 'Export failed', str(e))
